import os
import json
from flask import Blueprint, render_template, request, jsonify
import io
from datetime import datetime, date
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from flask import send_file

datos_compras_bp = Blueprint('datos_compras', __name__, url_prefix='/datos-compras')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.normpath(os.path.join(BASE_DIR, "..", "data"))
COMPRAS_FILE = os.path.join(DATA_DIR, "datoscompras.json")
RECEPCIONES_FILE = os.path.join(DATA_DIR, "recepciones_compras.json")

def _read_json(path):
    if not os.path.exists(path):
        return []
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)

def _write_json(path, data):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

def _to_int(value, default=0):
    try:
        if value is None or value == "":
            return default
        return int(value)
    except Exception:
        try:
            return int(float(value))
        except Exception:
            return default

def _to_float(value, default=0.0):
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default

def _normalizar_compra_payload(payload, compra_existente=None):
    """
    Normaliza los campos editables de una compra.
    Si compra_existente se envía, mantiene valores previos cuando no llegan en payload.
    """
    base = dict(compra_existente or {})

    text_fields = [
        "fecha_compra", "responsable_compra", "codigo", "nombre", "tipo_id",
        "identificacion", "tipo_empresa", "email", "descripcion", "tipo_compra",
        "pedido_proveedor", "factura_proveedor", "oc_coltrade", "estado_compra",
        "tipo_entrega", "fecha_llegada_nodo", "fecha_confirmacion_recoleccion",
        "proveedor", "tienda_compra", "palabra_clave_meli", "recepcion_en_odoo",
        "fecha_recibo", "banco_origen", "observaciones"
    ]
    numeric_float_fields = ["precio_unitario_sin_iva", "precio_unitario_con_iva", "valor_total_compra"]
    numeric_int_fields = ["cantidad_comprada"]

    for k in text_fields:
        if k in payload:
            v = payload.get(k)
            base[k] = "" if v is None else str(v).strip()

    for k in numeric_float_fields:
        if k in payload:
            base[k] = _to_float(payload.get(k), 0.0)

    for k in numeric_int_fields:
        if k in payload:
            base[k] = _to_int(payload.get(k), 0)

    # si no viene valor_total o viene en cero, recalcular
    precio = _to_float(base.get("precio_unitario_sin_iva", 0), 0.0)
    cantidad = _to_int(base.get("cantidad_comprada", 0), 0)
    valor_total = _to_float(base.get("valor_total_compra", 0), 0.0)
    if valor_total == 0:
        base["valor_total_compra"] = precio * cantidad

    return base

@datos_compras_bp.route('/')
def page():
    return render_template('datos_compras.html')

# GET: devuelve compras con estado cruzado con recepciones
@datos_compras_bp.route('/api', methods=['GET'])
def get_compras():
    compras = _read_json(COMPRAS_FILE)
    recepciones = _read_json(RECEPCIONES_FILE)
    recepciones_dict = {
        r.get("id_compra"): r
        for r in recepciones
        if isinstance(r, dict) and r.get("id_compra") is not None
    }
    compras_con_estado = []

    for compra in compras:
        id_compra = compra.get("id")
        cantidad = max(0, _to_int(compra.get("cantidad_comprada", 0), 0))
        recepcion = recepciones_dict.get(id_compra)

        if recepcion:
            unidades_recibidas = max(
                0,
                _to_int(
                    recepcion.get("unidades_recibidas", compra.get("unidades_recibidas", 0)),
                    0
                )
            )
            if cantidad > 0:
                unidades_recibidas = min(unidades_recibidas, cantidad)
            unidades_faltantes = max(0, cantidad - unidades_recibidas)

            if unidades_recibidas == 0:
                estado = "Pendiente"
            elif unidades_faltantes > 0:
                estado = "Parcial"
            else:
                estado = "Completa"

            rc_odoo = recepcion.get("rc_odoo", compra.get("recepcion_en_odoo", ""))
            if isinstance(rc_odoo, bool):
                rc_odoo = "Si" if rc_odoo else "No"
            elif rc_odoo is None:
                rc_odoo = ""

            fecha_recibo = (
                recepcion.get("fecha_recibo_odoo")
                or recepcion.get("fecha_recibo")
                or compra.get("fecha_recibo", "")
            )
            observaciones_ops = recepcion.get(
                "observaciones_ops",
                compra.get("observaciones_ops", "")
            )

            compra_actualizada = {
                **compra,
                "recepcion_en_odoo": str(rc_odoo).strip(),
                "fecha_recibo": fecha_recibo,
                "observaciones_ops": observaciones_ops,
                "unidades_recibidas": unidades_recibidas,
                "unidades_faltantes": unidades_faltantes,
                "estado_recepcion": estado
            }
        else:
            compra_actualizada = {
                **compra,
                "recepcion_en_odoo": compra.get("recepcion_en_odoo", ""),
                "fecha_recibo": compra.get("fecha_recibo", ""),
                "observaciones_ops": compra.get("observaciones_ops", ""),
                "unidades_recibidas": 0,
                "unidades_faltantes": cantidad,
                "estado_recepcion": "Pendiente"
            }

        compras_con_estado.append(compra_actualizada)

    return jsonify(compras_con_estado)

# POST: crear nueva compra (todos los campos que diligencias)
@datos_compras_bp.route('/api', methods=['POST'])
def crear_compra():
    payload = request.get_json()
    if not payload:
        return jsonify({"msg": "payload vacio"}), 400

    compras = _read_json(COMPRAS_FILE)
    new_id = max([c.get("id", 0) for c in compras], default=0) + 1

    # calcular valor_total_compra si no viene
    precio = float(payload.get("precio_unitario_sin_iva", 0))
    cantidad = int(payload.get("cantidad_comprada", 0))
    payload["valor_total_compra"] = precio * cantidad

    nuevo = {
        "id": new_id,
        **payload,
        # campos de recepcion por defecto
        "unidades_recibidas": 0,
        "unidades_faltantes": cantidad,
        "estado_recepcion": "Pendiente"
    }

    compras.append(nuevo)
    _write_json(COMPRAS_FILE, compras)
    return jsonify({"msg": "creado", "id": new_id})


@datos_compras_bp.route('/api/<int:compra_id>', methods=['PUT'])
def actualizar_compra(compra_id):
    payload = request.get_json()
    if not payload:
        return jsonify({"msg": "payload vacio"}), 400

    compras = _read_json(COMPRAS_FILE)
    idx = next((i for i, c in enumerate(compras) if c.get("id") == compra_id), -1)
    if idx == -1:
        return jsonify({"msg": "compra no encontrada"}), 404

    compra_actual = compras[idx]
    compra_actualizada = _normalizar_compra_payload(payload, compra_actual)
    compra_actualizada["id"] = compra_id

    # sincronizar estado de recepcion con datos de recepciones (si existe)
    recepciones = _read_json(RECEPCIONES_FILE)
    recep = next((r for r in recepciones if r.get("id_compra") == compra_id), None)

    cantidad = _to_int(compra_actualizada.get("cantidad_comprada", 0), 0)
    if recep:
        unidades_recibidas = _to_int(recep.get("unidades_recibidas", 0), 0)
        recep["cantidad_compra"] = cantidad
        recep["unidades_faltantes"] = max(0, cantidad - unidades_recibidas)
        compra_actualizada["unidades_recibidas"] = unidades_recibidas
        compra_actualizada["unidades_faltantes"] = recep["unidades_faltantes"]
    else:
        unidades_recibidas = _to_int(compra_actualizada.get("unidades_recibidas", 0), 0)
        compra_actualizada["unidades_recibidas"] = unidades_recibidas
        compra_actualizada["unidades_faltantes"] = max(0, cantidad - unidades_recibidas)

    if cantidad > 0 and compra_actualizada["unidades_recibidas"] >= cantidad:
        compra_actualizada["estado_recepcion"] = "Completa"
    elif compra_actualizada["unidades_recibidas"] == 0:
        compra_actualizada["estado_recepcion"] = "Pendiente"
    else:
        compra_actualizada["estado_recepcion"] = "Parcial"

    compras[idx] = compra_actualizada
    _write_json(COMPRAS_FILE, compras)
    _write_json(RECEPCIONES_FILE, recepciones)

    return jsonify({"msg": "actualizado", "id": compra_id})


@datos_compras_bp.route('/api/<int:compra_id>', methods=['DELETE'])
def eliminar_compra(compra_id):
    compras = _read_json(COMPRAS_FILE)
    compra_existente = next((c for c in compras if c.get("id") == compra_id), None)
    if not compra_existente:
        return jsonify({"msg": "compra no encontrada"}), 404

    compras_filtradas = [c for c in compras if c.get("id") != compra_id]
    _write_json(COMPRAS_FILE, compras_filtradas)

    recepciones = _read_json(RECEPCIONES_FILE)
    recepciones_filtradas = [r for r in recepciones if r.get("id_compra") != compra_id]
    deleted_recepciones = len(recepciones) - len(recepciones_filtradas)
    if deleted_recepciones > 0:
        _write_json(RECEPCIONES_FILE, recepciones_filtradas)

    return jsonify({
        "msg": "eliminado",
        "id": compra_id,
        "recepciones_eliminadas": deleted_recepciones
    })


# -----------------------
#  PLANTILLA EXCEL (descarga)
# -----------------------
@datos_compras_bp.route('/template.xlsx', methods=['GET'])
def download_template():
    """
    Genera y devuelve una plantilla Excel con cabeceras.
    El usuario puede abrirla, rellenar filas y subirla vía /import.
    """
    headers = [
        "fecha_compra",
        "responsable_compra",
        "codigo",
        "nombre",
        "tipo_id",
        "identificacion",
        "tipo_empresa",
        "email",
        "precio_unitario_sin_iva",
        "precio_unitario_con_iva",
        "cantidad_comprada",
        "valor_total_compra",  # opcional: se calculará si está vacío
        "descripcion",
        "tipo_compra",
        "pedido_proveedor",
        "factura_proveedor",
        "oc_coltrade",
        "estado_compra",
        "tipo_entrega",
        "fecha_llegada_nodo",
        "fecha_confirmacion_recoleccion",
        "proveedor",
        "tienda_compra",
        "palabra_clave_meli",
        "recepcion_en_odoo",
        "fecha_recibo",
        "banco_origen",
        "observaciones"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "plantilla_datos_compras"

    # escribir cabeceras
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(h) + 2)

    # fila de ejemplo (opcional, para guiar al usuario)
    example = [
        datetime.now().date().isoformat(),
        "Nombre Ejemplo",
        "ABC123",
        "Nombre cliente",
        "CC",
        "12345678",
        "Distribuidor",
        "cliente@ejemplo.com",
        1000,
        1190,
        10,
        "",  # valor_total_compra vacío -> se calculará
        "Descripción ejemplo",
        "Normal",
        "PO-111",
        "FAC-111",
        "OC-111",
        "Pendiente",
        "Domicilio",
        "",
        "",
        "Proveedor A",
        "Tienda 1",
        "meli-key",
        "No",
        "",
        "Banco X",
        "Observaciones..."
    ]
    for ci, v in enumerate(example, start=1):
        ws.cell(row=2, column=ci, value=v)

    # preparar buffer y devolver
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="plantilla_datos_compras.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# -----------------------
#  IMPORTAR EXCEL (subida)
# -----------------------
@datos_compras_bp.route('/import', methods=['POST'])
def import_from_excel():
    """
    Espera un multipart/form-data con campo 'file' (archivo .xlsx).
    Valida filas, crea nuevos IDs y añade al JSON existente.
    Retorna resumen con cantidad importada y errores por fila.
    """
    file = request.files.get('file')
    if not file:
        return jsonify({"msg": "No se recibió archivo 'file'"}), 400

    try:
        wb = load_workbook(filename=file, data_only=True)
    except Exception as e:
        return jsonify({"msg": "Error al leer el archivo Excel", "error": str(e)}), 400

    ws = wb.active

    # leer encabezados (fila 1)
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value is not None else "")

    # Campos permitidos / esperados (se mapearán si están)
    allowed = set([
        "fecha_compra", "responsable_compra", "codigo", "nombre", "tipo_id",
        "identificacion", "tipo_empresa", "email",
        "precio_unitario_sin_iva", "precio_unitario_con_iva",
        "cantidad_comprada", "valor_total_compra",
        "descripcion", "tipo_compra", "pedido_proveedor", "factura_proveedor",
        "oc_coltrade", "estado_compra", "tipo_entrega",
        "fecha_llegada_nodo", "fecha_confirmacion_recoleccion",
        "proveedor", "tienda_compra", "palabra_clave_meli",
        "recepcion_en_odoo", "fecha_recibo", "banco_origen", "observaciones"
    ])

    # índices de columnas que se usarán
    header_index = {h: i for i, h in enumerate(headers)}

    compras = _read_json(COMPRAS_FILE)
    next_id = max([c.get("id", 0) for c in compras], default=0) + 1

    imported = 0
    errors = []
    new_rows = []

    # iterar filas desde fila 2
    for row_idx in range(2, ws.max_row + 1):
        row_vals = {}
        empty_row = True
        for h, col_idx in header_index.items():
            if h not in allowed:
                continue
            cell = ws[row_idx][col_idx]
            val = cell.value
            if val is not None and val != "":
                empty_row = False
            row_vals[h] = val

        if empty_row:
            continue  # saltar filas vacías

        # validaciones simples
        try:
            cantidad = row_vals.get("cantidad_comprada", 0) or 0
            precio = row_vals.get("precio_unitario_sin_iva", 0) or 0

            # si la celda viene como fecha (openpyxl date), convertir a ISO string
            fc = row_vals.get("fecha_compra")
            if isinstance(fc, (datetime, date)):
                row_vals["fecha_compra"] = fc.isoformat()
            elif fc is None:
                row_vals["fecha_compra"] = ""

            # normalizar strings
            for k in list(row_vals.keys()):
                if isinstance(row_vals[k], str):
                    row_vals[k] = row_vals[k].strip()

            # convertir tipos
            try:
                cantidad_num = int(cantidad)
            except Exception:
                try:
                    cantidad_num = int(float(cantidad))
                except Exception:
                    cantidad_num = 0

            try:
                precio_num = float(precio or 0)
            except Exception:
                precio_num = 0.0

            # calcular valor_total_compra si está vacío o cero
            valor_total = row_vals.get("valor_total_compra") or 0
            try:
                valor_total_num = float(valor_total)
            except Exception:
                valor_total_num = 0.0

            if not valor_total_num or valor_total_num == 0:
                valor_total_num = precio_num * cantidad_num

            nuevo = {
                "id": next_id,
                **{k: (v if v is not None else "") for k, v in row_vals.items()},
                "precio_unitario_sin_iva": precio_num,
                "cantidad_comprada": cantidad_num,
                "valor_total_compra": valor_total_num,
                # campos por defecto de recepcion
                "unidades_recibidas": 0,
                "unidades_faltantes": cantidad_num,
                "estado_recepcion": "Pendiente"
            }

            compras.append(nuevo)
            new_rows.append({"row": row_idx, "id": next_id})
            next_id += 1
            imported += 1
        except Exception as e:
            errors.append({"row": row_idx, "error": str(e)})

    # persistir
    try:
        _write_json(COMPRAS_FILE, compras)
    except Exception as e:
        return jsonify({"msg": "Error al guardar datos", "error": str(e)}), 500

    return jsonify({
        "msg": "import finished",
        "imported": imported,
        "new_rows": new_rows,
        "errors": errors
    })

# -----------------------
#  EXPORTAR EXCEL (descarga)
# -----------------------
@datos_compras_bp.route('/export', methods=['GET'])
def export_to_excel():
    """
    Exporta todos los registros actuales a un archivo Excel y lo devuelve.
    - Si quieres filtrar por algún parámetro, puedes añadir query params y aplicar filtros.
    """
    compras = _read_json(COMPRAS_FILE)

    headers = [
        "id",
        "fecha_compra",
        "responsable_compra",
        "codigo",
        "nombre",
        "tipo_id",
        "identificacion",
        "tipo_empresa",
        "email",
        "precio_unitario_sin_iva",
        "precio_unitario_con_iva",
        "cantidad_comprada",
        "valor_total_compra",
        "descripcion",
        "tipo_compra",
        "pedido_proveedor",
        "factura_proveedor",
        "oc_coltrade",
        "estado_compra",
        "tipo_entrega",
        "fecha_llegada_nodo",
        "fecha_confirmacion_recoleccion",
        "proveedor",
        "tienda_compra",
        "palabra_clave_meli",
        "recepcion_en_odoo",
        "fecha_recibo",
        "banco_origen",
        "observaciones",
        "unidades_recibidas",
        "unidades_faltantes",
        "estado_recepcion"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "datos_compras_export"

    # headers
    for ci, h in enumerate(headers, start=1):
        ws.cell(row=1, column=ci, value=h)
        ws.column_dimensions[get_column_letter(ci)].width = max(12, len(h) + 2)

    # filas
    for ri, c in enumerate(compras, start=2):
        for ci, h in enumerate(headers, start=1):
            val = c.get(h, "")
            # asegurarse que las fechas sean strings legibles (ISO)
            if isinstance(val, (datetime, date)):
                val = val.isoformat()
            ws.cell(row=ri, column=ci, value=val)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"datos_compras_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# -----------------------
# EXPORTAR EXCEL - SOLO FALTANTES (parcial)
# -----------------------
@datos_compras_bp.route('/export/faltantes', methods=['GET'])
def export_faltantes_excel():
    """
    Exporta a Excel solo las compras cuyo estado de recepción sea 'Parcial'
    (derivado a partir de datos de recepciones o campo estado_recepcion).
    """
    # leer archivos
    compras = _read_json(COMPRAS_FILE)
    recepciones = _read_json(RECEPCIONES_FILE)
    recepciones_dict = {r.get("id_compra"): r for r in recepciones if r and r.get("id_compra") is not None}

    # función local para derivar estado al estilo del GET /api
    def derive_estado(compra):
        id_compra = compra.get("id")
        cantidad = int(compra.get("cantidad_comprada", 0) or 0)
        recep = recepciones_dict.get(id_compra)
        if recep:
            unidades_recibidas = int(recep.get("unidades_recibidas", 0) or 0)
            unidades_faltantes = cantidad - unidades_recibidas
            if unidades_recibidas == 0:
                return "Pendiente"
            elif unidades_faltantes > 0:
                return "Parcial"
            else:
                return "Completa"
        else:
            return "Pendiente"

    # filtrar compras que resulten 'Parcial'
    compras_parcial = []
    for c in compras:
        estado = (c.get("estado_recepcion") or "").strip()
        # si hay estado en el registro y es parcial usarlo (case-insensitive)
        if estado and estado.lower() == "parcial":
            compras_parcial.append(c)
            continue
        # sino derivar por recepciones
        derived = derive_estado(c)
        if derived.lower() == "parcial":
            compras_parcial.append(c)

    # preparar excel (mismas columnas que el export general)
    headers = [
        "id",
        "fecha_compra",
        "responsable_compra",
        "codigo",
        "nombre",
        "tipo_id",
        "identificacion",
        "tipo_empresa",
        "email",
        "precio_unitario_sin_iva",
        "precio_unitario_con_iva",
        "cantidad_comprada",
        "valor_total_compra",
        "descripcion",
        "tipo_compra",
        "pedido_proveedor",
        "factura_proveedor",
        "oc_coltrade",
        "estado_compra",
        "tipo_entrega",
        "fecha_llegada_nodo",
        "fecha_confirmacion_recoleccion",
        "proveedor",
        "tienda_compra",
        "palabra_clave_meli",
        "recepcion_en_odoo",
        "fecha_recibo",
        "banco_origen",
        "observaciones",
        "unidades_recibidas",
        "unidades_faltantes",
        "estado_recepcion"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "faltantes_parcial"

    # escribir cabeceras
    for ci, h in enumerate(headers, start=1):
        ws.cell(row=1, column=ci, value=h)
        ws.column_dimensions[get_column_letter(ci)].width = max(12, len(h) + 2)

    # escribir filas
    for ri, c in enumerate(compras_parcial, start=2):
        # intentar completar unidades_recibidas/unidades_faltantes con recepcion si existe
        recep = recepciones_dict.get(c.get("id"), {})
        unidades_recibidas = recep.get("unidades_recibidas") if recep.get("unidades_recibidas") is not None else c.get("unidades_recibidas", "")
        unidades_faltantes = recep.get("unidades_faltantes") if recep.get("unidades_faltantes") is not None else c.get("unidades_faltantes", "")
        row_obj = {
            **c,
            "unidades_recibidas": unidades_recibidas,
            "unidades_faltantes": unidades_faltantes
        }
        for ci, h in enumerate(headers, start=1):
            val = row_obj.get(h, "")
            if isinstance(val, (datetime, date)):
                val = val.isoformat()
            ws.cell(row=ri, column=ci, value=val)

    # si no hay filas, dejar una nota en la hoja
    if len(compras_parcial) == 0:
        ws.cell(row=2, column=1, value="No se encontraron compras con estado 'Parcial' (faltantes).")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"faltantes_parcial_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
